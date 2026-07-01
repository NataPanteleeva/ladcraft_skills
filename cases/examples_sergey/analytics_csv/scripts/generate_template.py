"""Generate sales_report_template.xlsx with 5 sheets and charts (dev-only)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

OUT = Path(__file__).resolve().parent.parent / "analytics_csv" / "templates" / "sales_report_template.xlsx"


def add_bar(ws, title: str, cat_min_row: int, cat_max_row: int, val_col: int, anchor: str = "D1"):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.y_axis.title = "Значение"
    cats = Reference(ws, min_col=1, min_row=cat_min_row, max_row=cat_max_row)
    data = Reference(ws, min_col=val_col, min_row=cat_min_row - 1, max_row=cat_max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 18
    ws.add_chart(chart, anchor)


def add_pie(ws, title: str, min_row: int, max_row: int, anchor: str = "E1"):
    chart = PieChart()
    chart.title = title
    labels = Reference(ws, min_col=1, min_row=min_row, max_row=max_row)
    data = Reference(ws, min_col=2, min_row=min_row - 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.height = 12
    chart.width = 16
    ws.add_chart(chart, anchor)


def add_line(ws, title: str, min_row: int, max_row: int, anchor: str = "E1"):
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Покупки"
    cats = Reference(ws, min_col=1, min_row=min_row, max_row=max_row)
    data = Reference(ws, min_col=2, min_row=min_row - 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 18
    ws.add_chart(chart, anchor)


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # Сводка
    ws = wb.create_sheet("Сводка", 0)
    ws.append(["Метрика", "Значение"])
    ws.append(["Число покупок (cart)", 0])
    ws.append(["Выручка, руб.", 0])
    ws.append(["Средний чек, руб.", 0])
    ws.append(["Уникальные покупатели", 0])
    add_bar(ws, "Сводка по продажам", 2, 5, 2)

    # Бренды
    ws = wb.create_sheet("Бренды", 1)
    ws.append(["Бренд", "Покупки", "Выручка, руб.", "Доля, %"])
    for _ in range(10):
        ws.append(["", 0, 0, 0])
    add_bar(ws, "Топ брендов по покупкам", 2, 11, 2)

    # Категории
    ws = wb.create_sheet("Категории", 2)
    ws.append(["Категория", "Покупки", "Выручка, руб.", "Доля, %"])
    for _ in range(10):
        ws.append(["", 0, 0, 0])
    add_pie(ws, "Категории по покупкам", 2, 11)

    # Воронка
    ws = wb.create_sheet("Воронка", 3)
    ws.append(["Этап", "Количество", "Конверсия, %"])
    ws.append(["Просмотры (view)", 0, 100])
    ws.append(["Покупки (purchase)", 0, 0])
    add_bar(ws, "Воронка продаж", 2, 3, 2)

    # Динамика
    ws = wb.create_sheet("Динамика", 4)
    ws.append(["Дата", "Покупки", "Выручка, руб."])
    for _ in range(31):
        ws.append(["", 0, 0])
    add_line(ws, "Динамика покупок", 2, 32)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Written: {OUT}")


if __name__ == "__main__":
    main()
