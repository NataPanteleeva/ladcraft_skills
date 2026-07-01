"""Shared helpers: CSV analytics, openpyxl report, R7 Disk client."""

import json
import re
import uuid
from datetime import datetime
from io import BytesIO
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

EMPTY_BRAND_LABEL = "(без бренда)"
EMPTY_CATEGORY_LABEL = "(без категории)"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def pick_string(primary: Any, fallback: Any = "") -> str:
    for candidate in (primary, fallback):
        if isinstance(candidate, dict):
            value = candidate.get("value")
            if isinstance(value, str) and value.strip():
                return value.strip()
        if isinstance(candidate, str) and candidate.strip():
            return candidate.strip()
    return ""


def read_user_env(state) -> dict:
    if state is None or not isinstance(state, dict):
        return {}
    env = state.get("environment") if isinstance(state.get("environment"), dict) else {}
    user = env.get("user") if isinstance(env.get("user"), dict) else {}
    flat = state.get("userEnv") if isinstance(state.get("userEnv"), dict) else {}
    out = {}
    out.update(flat)
    out.update(user)
    return out


def resolve_skill_storage(state: dict):
    caps = state.get("capabilities") if isinstance(state, dict) else None
    if not isinstance(caps, dict):
        return None
    raw = caps.get("skillStorage") or caps.get("storage") or caps.get("key-value-storage")
    if not isinstance(raw, dict):
        return None
    if not callable(raw.get("get")) or not callable(raw.get("set")):
        return None
    return raw


def resolve_positive_id(value: Any) -> int | None:
    if isinstance(value, dict):
        value = value.get("value")
    if isinstance(value, bool):
        return None
    if isinstance(value, int) and value > 0:
        return value
    if isinstance(value, float) and value > 0:
        return int(value)
    if isinstance(value, str) and value.strip():
        try:
            parsed = int(float(value.strip()))
            return parsed if parsed > 0 else None
        except ValueError:
            return None
    return None


def parse_csv_semicolon(text: str) -> dict:
    lines = [line for line in text.replace("\ufeff", "").splitlines() if line.strip()]
    if len(lines) < 2:
        return {"ok": False, "error": "CSV пуст или содержит только заголовок."}
    headers = [part.strip() for part in lines[0].split(";")]
    rows = []
    for line in lines[1:]:
        parts = line.split(";")
        row = {headers[i]: (parts[i] if i < len(parts) else "").strip() for i in range(len(headers))}
        rows.append(row)
    return {"ok": True, "headers": headers, "rows": rows}


def parse_price(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    normalized = str(value or "").strip().replace(" ", "").replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return 0.0


def parse_event_date(value: Any) -> str:
    raw = str(value or "").strip()
    match = re.match(r"^(\d{4}-\d{2}-\d{2})", raw)
    if match:
        return match.group(1)
    return raw[:10] if raw else "unknown"


def parse_event_month(value: Any) -> str:
    date = parse_event_date(value)
    if len(date) >= 7 and date[4] == "-":
        return date[:7]
    return date


MONTH_SHORT_RU = {
    "01": "янв",
    "02": "фев",
    "03": "мар",
    "04": "апр",
    "05": "май",
    "06": "июн",
    "07": "июл",
    "08": "авг",
    "09": "сен",
    "10": "окт",
    "11": "ноя",
    "12": "дек",
}


def format_month_label(year_month: str) -> str:
    parts = year_month.split("-")
    if len(parts) == 2 and len(parts[0]) == 4:
        month_label = MONTH_SHORT_RU.get(parts[1], parts[1])
        return f"{month_label} {parts[0]}"
    return year_month


def round2(num: float) -> float:
    return round(num, 2)


def filter_by_event_type(rows: list[dict], event_type: str) -> list[dict]:
    target = event_type.strip().lower()
    return [row for row in rows if str(row.get("event_type", "")).strip().lower() == target]


def build_top_list(items: list[dict], total_count: int, limit: int) -> list[dict]:
    sorted_items = sorted(items, key=lambda x: (-x["count"], -x["revenue"]))
    top = sorted_items[:limit]
    for item in top:
        item["share"] = round2((item["count"] / total_count) * 100) if total_count > 0 else 0.0
    return top


def aggregate_map(rows: list[dict], key_fn) -> list[dict]:
    bucket: dict[str, dict] = {}
    for row in rows:
        key = key_fn(row)
        price = parse_price(row.get("price"))
        entry = bucket.setdefault(key, {"key": key, "count": 0, "revenue": 0.0})
        entry["count"] += 1
        entry["revenue"] = round2(entry["revenue"] + price)
    return list(bucket.values())


def compute_analytics_from_rows(rows: list[dict]) -> dict:
    purchases = filter_by_event_type(rows, "cart")
    views = filter_by_event_type(rows, "view")
    purchase_events = filter_by_event_type(rows, "purchase")

    purchase_count = len(purchases)
    revenue = round2(sum(parse_price(row.get("price")) for row in purchases))
    unique_buyers = len({row.get("user_id") for row in purchases if row.get("user_id")})
    avg_check = round2(revenue / purchase_count) if purchase_count > 0 else 0.0

    view_count = len(views)
    funnel_purchase_count = len(purchase_events)
    funnel_conversion = round2((funnel_purchase_count / view_count) * 100) if view_count > 0 else 0.0

    brand_items = build_top_list(
        aggregate_map(purchases, lambda row: pick_string(row.get("brand")) or EMPTY_BRAND_LABEL),
        purchase_count,
        10,
    )
    brands = [
        {"brand": item["key"], "count": item["count"], "revenue": item["revenue"], "share": item["share"]}
        for item in brand_items
    ]

    def category_key(row: dict) -> str:
        code = pick_string(row.get("category_code"))
        return code.split(".")[0] if code else EMPTY_CATEGORY_LABEL

    category_items = build_top_list(aggregate_map(purchases, category_key), purchase_count, 10)
    categories = [
        {"category": item["key"], "count": item["count"], "revenue": item["revenue"], "share": item["share"]}
        for item in category_items
    ]

    dynamics_monthly_map: dict[str, dict] = {}
    for row in purchases:
        month_key = parse_event_month(row.get("event_time"))
        price = parse_price(row.get("price"))
        month_entry = dynamics_monthly_map.setdefault(
            month_key,
            {"month": month_key, "monthLabel": format_month_label(month_key), "count": 0, "revenue": 0.0},
        )
        month_entry["count"] += 1
        month_entry["revenue"] = round2(month_entry["revenue"] + price)
    dynamics_monthly = sorted(dynamics_monthly_map.values(), key=lambda x: x["month"])

    return {
        "summary": {
            "purchaseCount": purchase_count,
            "revenue": revenue,
            "avgCheck": avg_check,
            "uniqueBuyers": unique_buyers,
            "viewCount": view_count,
            "funnelPurchaseCount": funnel_purchase_count,
        },
        "brands": brands,
        "categories": categories,
        "funnel": [
            {"stage": "Просмотры (view)", "count": view_count, "conversion": 100.0},
            {"stage": "Покупки (purchase)", "count": funnel_purchase_count, "conversion": funnel_conversion},
        ],
        "dynamicsMonthly": dynamics_monthly,
    }


def analyze_csv_text(csv_text: str) -> dict:
    parsed = parse_csv_semicolon(csv_text)
    if not parsed.get("ok"):
        return parsed
    headers = parsed["headers"]
    for field in ("event_time", "event_type", "price"):
        if field not in headers:
            return {"ok": False, "error": f"В CSV отсутствует обязательная колонка «{field}»."}
    analytics = compute_analytics_from_rows(parsed["rows"])
    return {"ok": True, "analytics": analytics, "rowCount": len(parsed["rows"])}


RUBLE_NUMBER_FORMAT = r'# ##0,00\ "₽"'


def _apply_ruble_format(ws, col: int, row_start: int, row_end: int) -> None:
    if row_end < row_start:
        return
    for row in range(row_start, row_end + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = RUBLE_NUMBER_FORMAT


def _fix_xlsx_for_r7_office(xlsx_bytes: bytes) -> bytes:
    import zipfile

    inp = BytesIO(xlsx_bytes)
    out = BytesIO()
    with zipfile.ZipFile(inp, "r") as zin, zipfile.ZipFile(out, "w") as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "docProps/app.xml":
                text = data.decode("utf-8")
                text = text.replace("Microsoft Excel Compatible / Openpyxl", "Microsoft Excel")
                data = text.encode("utf-8")
            elif info.filename == "xl/styles.xml":
                text = data.decode("utf-8")
                text = text.replace(
                    "<fill><patternFill/></fill>",
                    '<fill><patternFill patternType="none"/></fill>',
                )
                data = text.encode("utf-8")
            zout.writestr(info, data)
    return out.getvalue()


def build_sales_report_xlsx(analytics: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList

    summary = analytics["summary"]
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Сводка", 0)
    ws.append(["Метрика", "Значение"])
    ws.append(["Число покупок (cart)", summary["purchaseCount"]])
    ws.append(["Выручка, руб.", summary["revenue"]])
    ws.append(["Средний чек, руб.", summary["avgCheck"]])
    ws.append(["Уникальные покупатели", summary["uniqueBuyers"]])
    _apply_ruble_format(ws, 2, 3, 4)

    ws = wb.create_sheet("Бренды", 1)
    ws.append(["Бренд", "Покупки", "Выручка, руб.", "Доля, %"])
    brands = analytics["brands"][:8]
    for item in brands:
        ws.append([item["brand"], item["count"], item["revenue"], item["share"]])
    if brands:
        last = 1 + len(brands)
        _apply_ruble_format(ws, 3, 2, last)
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "clustered"
        chart.title = "Топ брендов: покупки"
        chart.legend = None
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        ws.add_chart(chart, f"F{max(last + 2, 8)}")

    ws = wb.create_sheet("Категории", 2)
    ws.append(["Категория", "Покупки", "Выручка, руб.", "Доля, %"])
    categories = analytics["categories"][:8]
    for item in categories:
        ws.append([item["category"], item["count"], item["revenue"], item["share"]])
    if categories:
        last = 1 + len(categories)
        _apply_ruble_format(ws, 3, 2, last)
        pie = PieChart()
        pie.title = "Доли категорий"
        pie.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showCatName = True
        pie.dataLabels.showPercent = True
        ws.add_chart(pie, f"F{max(last + 2, 8)}")

    ws = wb.create_sheet("Воронка", 3)
    ws.append(["Этап", "Количество", "Конверсия, %"])
    funnel = analytics["funnel"]
    for step in funnel:
        ws.append([step["stage"], step["count"], step["conversion"]])
    if funnel:
        last = 1 + len(funnel)
        bar = BarChart()
        bar.type = "bar"
        bar.grouping = "clustered"
        bar.title = "Воронка view → purchase"
        bar.legend = None
        bar.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        ws.add_chart(bar, f"E{max(last + 2, 8)}")

    ws = wb.create_sheet("Динамика", 4)
    ws.append(["Месяц", "Покупки", "Выручка, руб."])
    monthly = analytics.get("dynamicsMonthly") or []
    for item in monthly:
        ws.append([item.get("monthLabel") or item.get("month"), item["count"], item["revenue"]])
    if monthly:
        last = 1 + len(monthly)
        _apply_ruble_format(ws, 3, 2, last)
        if len(monthly) >= 2:
            line = LineChart()
            line.title = "Динамика покупок по месяцам"
            line.legend = None
            line.add_data(Reference(ws, min_col=2, min_row=1, max_row=last), titles_from_data=True)
            line.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
            ws.add_chart(line, f"F{max(last + 2, 8)}")

    buffer = BytesIO()
    wb.save(buffer)
    return _fix_xlsx_for_r7_office(buffer.getvalue())


def list_report_sheets(analytics: dict) -> list[dict]:
    return [
        {"name": "Сводка", "table_rows": 5},
        {"name": "Бренды", "table_rows": len(analytics["brands"]) + 1},
        {"name": "Категории", "table_rows": len(analytics["categories"]) + 1},
        {"name": "Воронка", "table_rows": len(analytics["funnel"]) + 1},
        {"name": "Динамика", "table_rows": len(analytics.get("dynamicsMonthly") or []) + 1},
    ]


def build_default_output_name() -> str:
    return "отчет_продаж.xlsx"


def ensure_xlsx_extension(name: str) -> str:
    trimmed = str(name or "").strip()
    return trimmed if trimmed.lower().endswith(".xlsx") else f"{trimmed}.xlsx"


def stamp_output_name(name: str) -> str:
    dot = name.rfind(".")
    base = name[:dot] if dot > 0 else name
    ext = name[dot:] if dot > 0 else ".xlsx"
    return f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"


def _http_request(method: str, url: str, headers: dict | None = None, body: bytes | None = None) -> tuple[int, str]:
    req = Request(url, data=body, method=method)
    for key, value in (headers or {}).items():
        req.add_header(key, value)
    try:
        with urlopen(req, timeout=120) as response:
            return response.status, response.read().decode("utf-8", errors="replace")
    except HTTPError as err:
        return err.code, err.read().decode("utf-8", errors="replace")
    except URLError as err:
        raise RuntimeError(f"Сетевая ошибка {method} {url}: {err}") from err


def unwrap_api_data(payload: Any) -> Any:
    if isinstance(payload, dict) and "Response" in payload:
        response = payload.get("Response")
        if isinstance(response, dict) and "Data" in response:
            return response.get("Data", payload)
    return payload


def extract_document_id(payload: Any) -> int | None:
    if isinstance(payload, int):
        return payload
    data = unwrap_api_data(payload)
    if isinstance(data, int):
        return data
    if isinstance(data, dict):
        for key in ("Id", "id", "DocumentId"):
            value = data.get(key)
            if isinstance(value, int):
                return value
    return None


def resolve_base_url(state: dict, params: dict, skill_storage) -> str:
    from_param = pick_string(params.get("base_url"))
    if from_param:
        return from_param.rstrip("/")
    if skill_storage:
        cached = skill_storage.get("r7_disk_base_url")
        if isinstance(cached, str) and cached.strip():
            return cached.strip().rstrip("/")
    return pick_string(read_user_env(state).get("R7_DISK_BASE_URL")).rstrip("/")


async def ensure_auth_token(base_url: str, login: str, password: str, skill_storage, auth_token_param: Any) -> dict:
    auth_token = pick_string(auth_token_param)
    if not auth_token and skill_storage:
        cached = skill_storage.get("r7_disk_auth_token")
        if isinstance(cached, str) and cached.strip():
            auth_token = cached.strip()
    if auth_token:
        return {"ok": True, "auth_token": auth_token}
    if not login or not password:
        return {"ok": False, "error": "Нет auth_token и не заданы R7_DISK_LOGIN/R7_DISK_PASSWORD."}
    status, raw = _http_request(
        "POST",
        f"{base_url}/api/v2/auth/Login",
        headers={"Content-Type": "application/json; charset=utf-8"},
        body=json.dumps({"Login": login, "Password": password}).encode("utf-8"),
    )
    try:
        payload = json.loads(raw) if raw else {}
    except json.JSONDecodeError:
        return {"ok": False, "error": f"Login: ответ не JSON (HTTP {status})."}
    if status >= 400:
        return {"ok": False, "error": f"Login HTTP {status}: {raw[:240]}"}
    token = (
        payload.get("Response", {})
        .get("Data", {})
        .get("Tokens", {})
        .get("AuthToken")
    )
    if not isinstance(token, str) or not token:
        return {"ok": False, "error": "Login: AuthToken не найден."}
    if skill_storage:
        skill_storage.set("r7_disk_auth_token", token)
        skill_storage.set("r7_disk_base_url", base_url)
    return {"ok": True, "auth_token": token}


async def get_document_id_by_name(base_url: str, auth_token: str, directory_id: int, name: str) -> dict:
    url = (
        f"{base_url}/api/v1/Documents/GetIdByName?"
        f"name={_url_quote(name)}&directoryId={directory_id}"
    )
    status, raw = _http_request("GET", url, headers={"Authorization": auth_token})
    if status >= 400:
        return {"ok": False, "error": f"GetIdByName HTTP {status}: {raw[:240]}"}
    try:
        payload = json.loads(raw) if raw else {}
    except json.JSONDecodeError:
        payload = raw
    return {"ok": True, "documentId": extract_document_id(payload)}


async def delete_document(base_url: str, auth_token: str, document_id: int) -> dict:
    status, raw = _http_request(
        "POST",
        f"{base_url}/api/v1/Documents/Delete",
        headers={"Authorization": auth_token, "Content-Type": "application/json; charset=utf-8"},
        body=json.dumps({"Ids": [document_id]}).encode("utf-8"),
    )
    if status >= 400:
        return {"ok": False, "error": f"Delete HTTP {status}: {raw[:240]}"}
    return {"ok": True}


async def download_document_bytes(base_url: str, auth_token: str, document_id: int) -> dict:
    url = f"{base_url}/api/v1/Documents/Download?id={document_id}"
    req = Request(url, method="GET", headers={"Authorization": auth_token})
    try:
        with urlopen(req, timeout=120) as response:
            if response.status >= 400:
                return {"ok": False, "error": f"Download HTTP {response.status}"}
            return {"ok": True, "bytes": response.read()}
    except HTTPError as err:
        return {"ok": False, "error": f"Download HTTP {err.code}: {err.read()[:240]!r}"}
    except URLError as err:
        return {"ok": False, "error": f"Сетевая ошибка download: {err}"}


def _url_quote(value: str) -> str:
    from urllib.parse import quote

    return quote(value, safe="")


def _ascii_fallback_filename(file_name: str) -> str:
    dot = file_name.rfind(".")
    ext = file_name[dot:] if dot > 0 else ""
    base = (file_name[:dot] if dot > 0 else file_name).encode("ascii", "ignore").decode("ascii")
    base = re.sub(r"_+", "_", base).strip("_") or "file"
    return (base[:80] + ext)[:120]


def build_multipart_upload_body(boundary: str, file_name: str, content_type: str, file_bytes: bytes) -> bytes:
    ascii_name = _ascii_fallback_filename(file_name)
    utf8_name = _url_quote(file_name)
    preamble = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="{ascii_name}"; filename*=UTF-8\'\'{utf8_name}\r\n'
        f"Content-Type: {content_type}\r\n\r\n"
    ).encode("utf-8")
    epilogue = f"\r\n--{boundary}--\r\n".encode("utf-8")
    return preamble + file_bytes + epilogue


async def perform_multipart_upload(
    base_url: str,
    auth_token: str,
    directory_id: int,
    file_name: str,
    file_bytes: bytes,
    mime_override: str | None = None,
    *,
    replace_document_id: int | None = None,
) -> dict:
    boundary = f"----R7Disk{uuid.uuid4().hex}"
    content_type = mime_override or XLSX_MIME
    body = build_multipart_upload_body(boundary, file_name, content_type, file_bytes)
    headers = {
        "Authorization": auth_token,
        "DirectoryId": str(directory_id),
        "Content-Type": f"multipart/form-data; boundary={boundary}",
    }
    if replace_document_id is not None:
        headers["Id"] = str(replace_document_id)
        headers["DocumentId"] = str(replace_document_id)
    req = Request(
        f"{base_url}/api/v1/Documents/Upload",
        data=body,
        method="POST",
        headers=headers,
    )
    try:
        with urlopen(req, timeout=180) as response:
            raw = response.read().decode("utf-8", errors="replace")
            if response.status >= 400:
                return {"ok": False, "error": f"Upload HTTP {response.status}: {raw[:320]}"}
    except HTTPError as err:
        raw = err.read().decode("utf-8", errors="replace")
        return {"ok": False, "error": f"Upload HTTP {err.code}: {raw[:320]}"}
    except URLError as err:
        return {"ok": False, "error": f"Сетевая ошибка upload: {err}"}
    try:
        payload = json.loads(raw) if raw else {}
    except json.JSONDecodeError:
        payload = raw
    return {"ok": True, "document_id": extract_document_id(payload), "data": payload}


async def resolve_output_name_by_policy(
    base_url: str, auth_token: str, directory_id: int, name: str, policy: str
) -> dict:
    exists = await get_document_id_by_name(base_url, auth_token, directory_id, name)
    if not exists.get("ok"):
        return {"ok": True, "name": name, "existing_document_id": None}
    document_id = exists.get("documentId")
    if document_id is None:
        return {"ok": True, "name": name, "existing_document_id": None}
    if policy == "overwrite":
        return {"ok": True, "name": name, "existing_document_id": document_id}
    if policy == "error":
        return {"ok": False, "error": f"Файл «{name}» уже существует в папке id={directory_id}."}
    return {"ok": True, "name": stamp_output_name(name), "existing_document_id": None}


async def upload_replacing_document(
    base_url: str,
    auth_token: str,
    directory_id: int,
    file_name: str,
    file_bytes: bytes,
    mime_override: str | None,
    existing_document_id: int | None,
) -> dict:
    if existing_document_id is not None:
        in_place = await perform_multipart_upload(
            base_url,
            auth_token,
            directory_id,
            file_name,
            file_bytes,
            mime_override,
            replace_document_id=existing_document_id,
        )
        if in_place.get("ok"):
            return {
                **in_place,
                "document_id": in_place.get("document_id") or existing_document_id,
                "upload_method": "in_place_id_header",
            }
        deleted = await delete_document(base_url, auth_token, existing_document_id)
        if not deleted.get("ok"):
            return {
                "ok": False,
                "error": (
                    f"Не удалось обновить «{file_name}» (in-place и delete): "
                    f"{in_place.get('error')}; {deleted.get('error')}"
                ),
            }
    uploaded = await perform_multipart_upload(
        base_url,
        auth_token,
        directory_id,
        file_name,
        file_bytes,
        mime_override,
    )
    if uploaded.get("ok"):
        uploaded["upload_method"] = "new_upload"
    return uploaded
